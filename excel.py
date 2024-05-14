import pandas as pd
import openpyxl
from parser.pars import groups_list, users
import datetime

import logging


logger = logging.getLogger(__name__)

file_handler = logging.FileHandler('logs.txt', encoding='utf8')
file_handler.setFormatter(logging.Formatter(
    fmt='[%(asctime)s] #%(levelname)-8s %(name)s '
           '%(funcName)s:%(lineno)d - %(message)s'))
logger.addHandler(file_handler)



today = (datetime.datetime.today() + datetime.timedelta(days=0)).strftime('%Y-%m-%d')

def month(month):
    if month < 12:
        return list(pd.period_range(start=f'2024-{month}-01', end=f'2024-{month+1}-01', freq='D'))[:-1]
    return list(pd.period_range(start=f'2024-{month}-01', end=f'2025-01-01', freq='D'))[:-1]

lexicon_month = {'01':'Январь',
                 '02':'Февраль',
                 '03':'Март',
                 '04':'Апрель',
                 '05':'Май',
                 '06':'Июнь',
                 '07':'Июль',
                 '08':'Август',
                 '09':'Сентябрь',
                 '10':'Октябрь',
                 '11':'Ноябрь',
                 '12':'Декабрь'}

def create_table(name):
    file_name = f'{name}.xlsx'

    month_1 = pd.DataFrame(index=groups_list)
    for day in month(1):
        month_1[day] = '-'
    month_1.to_excel(file_name, sheet_name=lexicon_month[1])

    with pd.ExcelWriter(file_name, mode='a', engine='openpyxl') as writer:
        for i in range(2,13):
            page = pd.DataFrame(index=groups_list)

            for day in month(i):
                page[day] = '-'

            page.to_excel(writer, sheet_name=lexicon_month[i])

def add_stat(day=today, file_name='2024'):
    logger.info('\n\n-----start add stats----- \n')


    workbook = openpyxl.load_workbook(f'{file_name}.xlsx') #открытие файла
    logger.info('open file :1')

    sheet = workbook[lexicon_month[day[5:7]]] #открываем нужный лист с текущим месяцем
    logger.info('open page :2')

    for col in range(1, sheet.max_column+1): #проходимся по всем колонкам в первой строке (даты)
        if sheet.cell(row=1, column=col).value == day: #если дата совпадает с текущей, сохраняем ее в переменную
            column_number = col
            logger.info('FIND DAY COLUMN :3')
            break

    count_users = users()
    logger.info(f'{"-"*15}\n\nсловарь со статистикой - \nдлина: {len(count_users)}\nсловарь: {count_users}\n\n{"-"*15}')

    for row in range(2, sheet.max_row + 1): #проходимся по всем группам (первый столбец)
        try:
            group = (sheet.cell(row=row, column=1).value)
            if group in count_users: # если номер группы есть в словаре со статистикой, добавляем в таблицу количество пользователей в этой группе
                sheet.cell(row = row, column=column_number).value = '-'
                sheet.cell(row=row, column=column_number).value = str(count_users[group])
        except Exception as error:
            logger.info(f'!!!{error}!!!')

    sheet.cell(row=1, column=1).value = f'{datetime.datetime.today():%d.%m.%Y}' # добавляем в левый верхний угол дату изменения (для выявления ошибок)

    try:
        workbook.save(f'{file_name}.xlsx')
        logger.info('save new stats :4\n\n') # сохраняем изменения
    except Exception as x:
        logger.error(f'save new stats :6 - [{x}]\n\n')
